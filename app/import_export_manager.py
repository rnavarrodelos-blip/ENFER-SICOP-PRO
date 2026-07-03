from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import csv
from app.db_init import get_session
from app.models import Participante, Curso, Voucher, Certificado
from app.link_integration import LinkIntegration
from app.auditoria_manager import AuditoriaManager
from datetime import datetime


class ImportExportManager:
    """Gestiona importación/exportación de datos"""

    def __init__(self):
        self.session = get_session()
        self.link = LinkIntegration()
        self.auditoria = AuditoriaManager()

    def importar_participantes_excel(self, archivo_path, curso_id):
        """
        Importa participantes desde Excel
        Formato esperado: Nombre | Email | Teléfono | Identificación
        """
        try:
            workbook = load_workbook(archivo_path)
            sheet = workbook.active
            resultados = {"exitosos": 0, "errores": 0, "duplicados": 0, "detalles": []}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay nombre
                    nombre = row[0]
                    email = row[1] if len(row) > 1 else None
                    telefono = row[2] if len(row) > 2 else None
                    identificacion = row[3] if len(row) > 3 else None

                    # Verificar duplicado
                    existente = self.session.query(Participante).filter_by(
                        email=email
                    ).first() if email else None

                    if existente:
                        resultados["duplicados"] += 1
                        resultados["detalles"].append(f"Duplicado: {nombre}")
                    else:
                        try:
                            nuevo_participante = Participante(
                                nombre=nombre,
                                email=email,
                                telefono=telefono,
                                identificacion=identificacion,
                                curso_id=curso_id
                            )
                            self.session.add(nuevo_participante)
                            self.session.commit()

                            # Registrar en auditoría
                            self.auditoria.registrar_cambio(
                                "participantes",
                                "CREATE",
                                nuevo_participante.id,
                                datos_nuevos={"nombre": nombre, "email": email}
                            )

                            resultados["exitosos"] += 1
                        except Exception as e:
                            resultados["errores"] += 1
                            resultados["detalles"].append(f"Error en {nombre}: {str(e)}")

            return resultados
        except Exception as e:
            return {"exitosos": 0, "errores": 1, "duplicados": 0, "detalles": [str(e)]}

    def importar_vouchers_excel(self, archivo_path):
        """
        Importa vouchers desde Excel
        Formato: Número Voucher | Email Participante | Monto
        """
        try:
            workbook = load_workbook(archivo_path)
            sheet = workbook.active
            resultados = {"exitosos": 0, "errores": 0, "duplicados": 0, "detalles": []}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay número de voucher
                    numero_voucher = str(row[0])
                    email_participante = row[1] if len(row) > 1 else None
                    monto = float(row[2]) if len(row) > 2 else 0.0

                    # Buscar participante
                    participante = self.session.query(Participante).filter_by(
                        email=email_participante
                    ).first()

                    if not participante:
                        resultados["errores"] += 1
                        resultados["detalles"].append(f"Participante no encontrado: {email_participante}")
                        continue

                    # Verificar duplicado
                    if not self.link.validar_numero_voucher(numero_voucher):
                        resultados["duplicados"] += 1
                        resultados["detalles"].append(f"Duplicado: {numero_voucher}")
                        continue

                    try:
                        nuevo_voucher = Voucher(
                            numero_voucher=numero_voucher,
                            participante_id=participante.id,
                            monto=monto,
                            estado="verificado",
                            verified=True
                        )
                        self.session.add(nuevo_voucher)
                        self.session.commit()

                        # Registrar en auditoría
                        self.auditoria.registrar_cambio(
                            "vouchers",
                            "CREATE",
                            nuevo_voucher.id,
                            datos_nuevos={"numero": numero_voucher, "monto": monto}
                        )

                        resultados["exitosos"] += 1
                    except Exception as e:
                        resultados["errores"] += 1
                        resultados["detalles"].append(f"Error en {numero_voucher}: {str(e)}")

            return resultados
        except Exception as e:
            return {"exitosos": 0, "errores": 1, "duplicados": 0, "detalles": [str(e)]}

    def exportar_reporte_excel(self, archivo_path, tipo="completo"):
        """
        Exporta un reporte a Excel
        tipo: completo, participantes, vouchers, certificados
        """
        try:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)

            if tipo == "completo" or tipo == "participantes":
                self._exportar_participantes(wb)
            if tipo == "completo" or tipo == "vouchers":
                self._exportar_vouchers(wb)
            if tipo == "completo" or tipo == "certificados":
                self._exportar_certificados(wb)

            wb.save(archivo_path)
            return {"success": True, "archivo": archivo_path}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _exportar_participantes(self, workbook):
        """Exporta participantes a hoja de Excel"""
        ws = workbook.create_sheet("Participantes")
        ws.append(["ID", "Nombre", "Email", "Teléfono", "Identificación", "Curso", "Estado", "Fecha Registro"])

        participantes = self.session.query(Participante).all()
        for participante in participantes:
            curso_nombre = participante.curso.nombre if participante.curso else "N/A"
            ws.append([
                participante.id,
                participante.nombre,
                participante.email,
                participante.telefono,
                participante.identificacion,
                curso_nombre,
                participante.estado,
                participante.created_at
            ])

    def _exportar_vouchers(self, workbook):
        """Exporta vouchers a hoja de Excel"""
        ws = workbook.create_sheet("Vouchers")
        ws.append(["ID", "Número", "Participante", "Monto", "Estado", "Verificado", "Fecha Emisión"])

        vouchers = self.session.query(Voucher).all()
        for voucher in vouchers:
            participante_nombre = voucher.participante.nombre if voucher.participante else "N/A"
            verified = "Sí" if voucher.verified else "No"
            ws.append([
                voucher.id,
                voucher.numero_voucher,
                participante_nombre,
                voucher.monto,
                voucher.estado,
                verified,
                voucher.fecha_emision
            ])

    def _exportar_certificados(self, workbook):
        """Exporta certificados a hoja de Excel"""
        ws = workbook.create_sheet("Certificados")
        ws.append(["ID", "Número", "Participante", "Curso", "Estado", "Fecha Emisión"])

        certificados = self.session.query(Certificado).all()
        for certificado in certificados:
            participante_nombre = certificado.participante.nombre if certificado.participante else "N/A"
            curso_nombre = certificado.curso.nombre if certificado.curso else "N/A"
            ws.append([
                certificado.id,
                certificado.numero_certificado,
                participante_nombre,
                curso_nombre,
                certificado.estado,
                certificado.fecha_emision
            ])
